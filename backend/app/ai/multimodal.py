"""Utilities for building multimodal content parts from attachment files.

Supported conversions:
- image/*       → base64 image_url data URI (analysed by vision-capable models)
- video/*       → up to 8 evenly-spaced frames extracted via OpenCV, each sent
                  as a base64 JPEG image_url part (reliable Gemini vision path)
- text/csv      → parsed text table injected as a text part
- Excel (.xls/xlsx) → parsed text table injected as a text part
- Everything else → skipped (handled by text metadata in the prompt)
"""

from __future__ import annotations

import base64
import csv
import io
import logging
from pathlib import Path
from typing import Any

logger = logging.getLogger(__name__)

# Maximum file size to encode inline (20 MB). Larger files are skipped.
_MAX_INLINE_BYTES = 20 * 1024 * 1024

# Number of frames to extract from a video for vision analysis
_VIDEO_MAX_FRAMES = 8


def _b64_data_uri(mime_type: str, data: bytes) -> str:
    encoded = base64.b64encode(data).decode("ascii")
    return f"data:{mime_type};base64,{encoded}"


def _extract_video_frames(path: Path) -> list[dict[str, Any]]:
    """Extract up to _VIDEO_MAX_FRAMES evenly-spaced frames from a video file.

    Returns a list of image_url content parts (JPEG-encoded frames).
    Falls back to an empty list if opencv is unavailable or the file is unreadable.
    """
    try:
        import cv2  # noqa: PLC0415
    except ImportError:
        logger.warning("opencv-python-headless not installed; video frame extraction unavailable")
        return []

    cap = cv2.VideoCapture(str(path))
    if not cap.isOpened():
        logger.warning("Could not open video file: %s", path)
        return []

    try:
        total_frames = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))
        if total_frames <= 0:
            return []

        # Pick _VIDEO_MAX_FRAMES evenly-spaced frame indices
        step = max(1, total_frames // _VIDEO_MAX_FRAMES)
        indices = list(range(0, total_frames, step))[:_VIDEO_MAX_FRAMES]

        parts: list[dict[str, Any]] = []
        for idx in indices:
            cap.set(cv2.CAP_PROP_POS_FRAMES, idx)
            ret, frame = cap.read()
            if not ret:
                continue
            success, buf = cv2.imencode(".jpg", frame, [cv2.IMWRITE_JPEG_QUALITY, 80])
            if not success:
                continue
            jpeg_bytes = buf.tobytes()
            parts.append({
                "type": "image_url",
                "image_url": {"url": _b64_data_uri("image/jpeg", jpeg_bytes)},
            })

        return parts
    finally:
        cap.release()


def _excel_to_text(data: bytes, filename: str) -> str:
    """Convert an Excel file to a plain-text table using openpyxl."""
    try:
        import openpyxl  # noqa: PLC0415

        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        parts: list[str] = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue
            col_count = max(len(r) for r in rows)
            lines: list[str] = [f"### Sheet: {sheet_name}"]
            for i, row in enumerate(rows):
                cells = [str(c) if c is not None else "" for c in row]
                cells += [""] * (col_count - len(cells))
                lines.append("| " + " | ".join(cells) + " |")
                if i == 0:
                    lines.append("| " + " | ".join(["---"] * col_count) + " |")
            parts.append("\n".join(lines))
        wb.close()
        return "\n\n".join(parts) if parts else "(empty workbook)"
    except Exception as exc:
        logger.warning("Failed to parse Excel file %s: %s", filename, exc)
        return f"(Could not parse Excel file: {exc})"


def _csv_to_text(data: bytes, filename: str) -> str:
    """Convert a CSV file to a plain-text markdown table."""
    try:
        text = data.decode("utf-8", errors="replace")
        reader = csv.reader(io.StringIO(text))
        rows = list(reader)
        if not rows:
            return "(empty CSV)"
        col_count = max(len(r) for r in rows)
        lines: list[str] = []
        for i, row in enumerate(rows):
            cells = row + [""] * (col_count - len(row))
            lines.append("| " + " | ".join(cells) + " |")
            if i == 0:
                lines.append("| " + " | ".join(["---"] * col_count) + " |")
        return "\n".join(lines)
    except Exception as exc:
        logger.warning("Failed to parse CSV file %s: %s", filename, exc)
        return f"(Could not parse CSV file: {exc})"


def build_multimodal_parts(
    attachments: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    """Return a list of LangChain-compatible content parts for the given attachments.

    Each attachment dict must include ``stored_path``, ``mime_type``, and ``filename``.
    """
    parts: list[dict[str, Any]] = []

    for att in attachments:
        mime_type: str = att.get("mime_type", "application/octet-stream")
        filename: str = att.get("filename", "file")
        stored_path: str | None = att.get("stored_path")

        if not stored_path:
            continue

        path = Path(stored_path)
        if not path.exists():
            logger.warning("Attachment file not found on disk: %s", stored_path)
            continue

        if mime_type.startswith("image/"):
            try:
                data = path.read_bytes()
            except OSError as exc:
                logger.warning("Could not read attachment %s: %s", stored_path, exc)
                continue
            if len(data) > _MAX_INLINE_BYTES:
                logger.info("Image %s too large for inline encoding, skipping", filename)
                continue
            parts.append({
                "type": "image_url",
                "image_url": {"url": _b64_data_uri(mime_type, data)},
            })

        elif mime_type.startswith("video/"):
            # Extract frames and send each as a JPEG image_url — reliable Gemini vision path
            frame_parts = _extract_video_frames(path)
            if frame_parts:
                parts.append({
                    "type": "text",
                    "text": f"The following {len(frame_parts)} frames were extracted from the video '{filename}'. Analyse them to answer the user's question:",
                })
                parts.extend(frame_parts)
            else:
                logger.warning("No frames extracted from video %s", filename)

        elif mime_type == "text/csv" or filename.lower().endswith(".csv"):
            try:
                data = path.read_bytes()
            except OSError as exc:
                logger.warning("Could not read attachment %s: %s", stored_path, exc)
                continue
            table_text = _csv_to_text(data, filename)
            parts.append({
                "type": "text",
                "text": f"Contents of {filename}:\n\n{table_text}",
            })

        elif (
            mime_type in {
                "application/vnd.ms-excel",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            }
            or filename.lower().endswith((".xls", ".xlsx"))
        ):
            try:
                data = path.read_bytes()
            except OSError as exc:
                logger.warning("Could not read attachment %s: %s", stored_path, exc)
                continue
            table_text = _excel_to_text(data, filename)
            parts.append({
                "type": "text",
                "text": f"Contents of {filename}:\n\n{table_text}",
            })

        # PDFs are handled via RAG; other types get metadata-only treatment.

    return parts
